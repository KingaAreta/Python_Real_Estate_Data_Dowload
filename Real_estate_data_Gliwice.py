#!/usr/bin/env python
# coding: utf-8

# In[2]:


import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from openpyxl import Workbook


# In[4]:


##Download data from moniter.pl##

gecko_service = Service(executable_path=GeckoDriverManager().install())
driver = webdriver.Firefox(service=gecko_service)

# visit the website
driver.get("https://moniter.pl/")

#log in

button = driver.find_element(By.XPATH, '/html/body/header/div[2]/nav/div/a[1]')
button.click()
email_field = driver.find_element(By.XPATH, '//*[@id="email"]')

email_input=str(input("Podaj e-mail do strony moniter: "))
password_input=str(input("Podaj hasło do strony moniter: "))

email_field.send_keys(email_input)
password_field = driver.find_element(By.XPATH, '//*[@id="password"]')
password_field.send_keys(password_input)

button = driver.find_element(By.XPATH, '/html/body/section/div/div/div[1]/div[2]/form/button')
button.click()

# find element "Gliwice"
link = driver.find_element(By.XPATH, '//a[@class="black-link" and contains(text(), "Gliwice")]')
driver.execute_script("arguments[0].click();", link)

time.sleep(5)

# download csv
link = driver.find_element(By.CLASS_NAME, 'export-csv')
driver.execute_script("arguments[0].click();", link)

time.sleep(5)
driver.quit()

# acquiring information about the new file

dir_path = r'C:\Users\Kinga\Downloads'

if os.path.exists(dir_path) and os.path.isdir(dir_path):
    files = os.listdir(dir_path)
    if files:
        newest_file = max(files, key=lambda f: os.path.getctime(os.path.join(dir_path, f)))
        print(f"Pobrany plik nazywa się: {newest_file}")

# name the file and location

        nazwa = newest_file  
        file_path = os.path.join(dir_path, nazwa)  

        if os.path.exists(file_path):
            file_creation_time = os.path.getctime(file_path)
            creation_date = datetime.fromtimestamp(file_creation_time).strftime('%Y-%m-%d')
            print(f'Data utworzenia pliku: {creation_date}')
        else:
            print('Plik nie istnieje.')
    else:
        print("Brak plików w katalogu.")
else:
    print("Podana ścieżka nie istnieje.")
print("Dane zostały pobrane")


# In[5]:


# Changing the file format


data_file_convert=str(input("Wpisz datę w formacie RRRR-MM-DD. Zostanie użyta do nazwania pliku: "))
downl_file=newest_file
correct_file_name=data_file_convert+"_gliwice"

# file locations
input_file = f'C:\\Users\\Kinga\\Downloads\\{downl_file}'
output_file = f'D:\\Nieruchomosci\\python_dane\\gliwice_moniter_dane\\{correct_file_name}.xlsx'


# eead a text file and extract data line by line
with open(input_file, 'r', encoding='utf-8-sig') as file:
    lines = file.readlines()

# create a new XLSX file
wb = Workbook()
ws = wb.active

# save header 
header = lines[0].strip().split(';')
for col_num, col_name in enumerate(header, 1):
    ws.cell(row=1, column=col_num).value = col_name

# save data from a CSV file to an XLSX file
for line_num, line in enumerate(lines[1:], 2):
    data = line.strip().split(';')
    for col_num, value in enumerate(data, 1):
        ws.cell(row=line_num, column=col_num).value = value

# save XLSX
wb.save(output_file)

print("Plik zapisany w ścieżce D:\\Nieruchomosci\\python_dane\\gliwice_moniter_dane\\")

